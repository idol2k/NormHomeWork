import csv
from openpyxl import Workbook

csv_path = 'data.csv'

rows = []
with open(csv_path, 'r', newline='') as csv_file:
    csv_reader = csv.reader(csv_file)
    header = next(csv_reader)
    for row in csv_reader:
        rows.append(row[:2] + row[3:])

print(rows)

excel_file = 'data.xlsx'
wb = Workbook()
ws = wb.active

for row_idx, item in enumerate(header[:3], start=2):
    ws.cell(row=row_idx, column=1, value=item)

for col_idx, i in enumerate(range(1, 7), start=2):
    ws.cell(row=1, column=col_idx, value='Person' + str(i))

for col_idx, item in enumerate(rows, start=2):
    for row_idx, value in enumerate(item, start=2):
        ws.cell(row=row_idx, column=col_idx, value=value)

ws.cell(row=4, column=1, value=header[3])

wb.save(excel_file)